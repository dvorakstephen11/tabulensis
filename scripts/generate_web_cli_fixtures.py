#!/usr/bin/env python3
import argparse
import os
import subprocess
import sys
from pathlib import Path


def write_workbook(path: Path, a1_value: str) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = a1_value
    ws["B1"] = "static"
    ws["A2"] = 1
    ws["B2"] = 2
    wb.save(path)


def run_cli(bin_path: str, old_path: Path, new_path: Path, fmt: str, output_path: Path) -> None:
    cmd = [bin_path, "diff", "--format", fmt, str(old_path), str(new_path)]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode not in (0, 1):
        sys.stderr.write(result.stderr)
        raise RuntimeError(f"excel-diff failed with exit code {result.returncode}")
    output_path.write_text(result.stdout, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate CLI payload/outcome fixtures for web UI tests.")
    parser.add_argument("--output-dir", type=Path, required=True, help="Directory for JSON fixtures.")
    parser.add_argument(
        "--bin",
        dest="bin_path",
        default=os.environ.get("EXCEL_DIFF_BIN"),
        help="Path to excel-diff binary (or set EXCEL_DIFF_BIN).",
    )
    args = parser.parse_args()

    if not args.bin_path:
        raise SystemExit("Missing --bin (or EXCEL_DIFF_BIN) for excel-diff.")

    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    old_path = output_dir / "web_fixture_old.xlsx"
    new_path = output_dir / "web_fixture_new.xlsx"
    payload_path = output_dir / "payload.json"
    outcome_path = output_dir / "outcome.json"

    write_workbook(old_path, "before")
    write_workbook(new_path, "after")

    run_cli(args.bin_path, old_path, new_path, "payload", payload_path)
    run_cli(args.bin_path, old_path, new_path, "outcome", outcome_path)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
