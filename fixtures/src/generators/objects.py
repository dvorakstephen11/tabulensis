import shutil
from pathlib import Path
from typing import Any, Dict, List, Union

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.workbook.defined_name import DefinedName

from .base import BaseGenerator


class NamedRangesGenerator(BaseGenerator):
    """Generates a workbook pair that exercises workbook- and sheet-scoped defined names."""

    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("named_ranges generator expects exactly two output filenames")

        sheet = self.args.get("sheet", "Sheet1")

        def create_workbook(global_names: List[DefinedName], local_name: DefinedName, path: Path):
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = sheet
            wb.create_sheet("Sheet2")

            ws1["A1"] = 1
            ws1["A2"] = 2
            ws1["A3"] = 3
            ws1["B1"] = 4
            ws1["C1"] = 5
            ws1["C2"] = 6
            ws1["D1"] = 7
            ws1["D2"] = 8

            for name in global_names:
                wb.defined_names.add(name)
            wb.defined_names.add(local_name)

            wb.save(path)

        global_keep = DefinedName("GlobalKeep", attr_text=f"{sheet}!$A$1:$A$3")
        global_remove = DefinedName("GlobalRemove", attr_text=f"{sheet}!$B$1")
        local_change_a = DefinedName("LocalChange", attr_text=f"{sheet}!$C$1", localSheetId=0)

        global_add = DefinedName("GlobalAdd", attr_text=f"{sheet}!$D$1:$D$2")
        local_change_b = DefinedName("LocalChange", attr_text=f"{sheet}!$C$2", localSheetId=0)

        create_workbook(
            global_names=[global_keep, global_remove],
            local_name=local_change_a,
            path=output_dir / output_names[0],
        )
        create_workbook(
            global_names=[global_keep, global_add],
            local_name=local_change_b,
            path=output_dir / output_names[1],
        )


class ChartsGenerator(BaseGenerator):
    """Generates a workbook pair that exercises chart add/remove/change detection."""

    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("charts generator expects exactly two output filenames")

        sheet = self.args.get("sheet", "Sheet1")

        def create_workbook(path: Path, chart1: Any, chart2: Any = None):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet

            ws.append(["X", "Y"])
            for idx in range(1, 6):
                ws.append([idx, idx * 2])

            data = Reference(ws, min_col=2, min_row=1, max_row=6)
            cats = Reference(ws, min_col=1, min_row=2, max_row=6)

            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            ws.add_chart(chart1, "E2")

            if chart2 is not None:
                chart2.add_data(data, titles_from_data=True)
                chart2.set_categories(cats)
                ws.add_chart(chart2, "E18")

            wb.save(path)

        create_workbook(output_dir / output_names[0], chart1=LineChart())
        create_workbook(output_dir / output_names[1], chart1=BarChart(), chart2=LineChart())


class CopyTemplateGenerator(BaseGenerator):
    """Copies a binary template file into the output directory."""

    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        template_arg = self.args.get("template")
        if not template_arg:
            raise ValueError("copy_template generator requires 'template' argument")

        template = Path(template_arg)
        if not template.exists():
            candidate = Path("fixtures") / template_arg
            if candidate.exists():
                template = candidate
            else:
                raise FileNotFoundError(f"Template {template} not found.")

        for name in output_names:
            shutil.copyfile(template, output_dir / name)
