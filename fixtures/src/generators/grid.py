import openpyxl
import zipfile
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Union, List, Dict, Any
from .base import BaseGenerator

class BasicGridGenerator(BaseGenerator):
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]
            
        rows = self.args.get('rows', 5)
        cols = self.args.get('cols', 5)
        two_sheets = self.args.get('two_sheets', False)
        
        for name in output_names:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Fill grid
            for r in range(1, rows + 1):
                for c in range(1, cols + 1):
                    ws.cell(row=r, column=c, value=f"R{r}C{c}")
            
            # Check if we need a second sheet
            if two_sheets:
                ws2 = wb.create_sheet(title="Sheet2")
                # Different dimensions for Sheet2 (PG1 requirement: 5x2)
                # If args are customized we might need more logic, but for PG1 this is sufficient or we use defaults
                s2_rows = 5
                s2_cols = 2
                for r in range(1, s2_rows + 1):
                    for c in range(1, s2_cols + 1):
                         ws2.cell(row=r, column=c, value=f"S2_R{r}C{c}")

            wb.save(output_dir / name)

class SparseGridGenerator(BaseGenerator):
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]
            
        for name in output_names:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sparse"
            
            # Specifics for pg1_sparse_used_range
            ws['A1'] = "A1"
            ws['B2'] = "B2"
            ws['G10'] = "G10" # Forces extent
            # Row 5 and Col D are empty implicitly by not writing to them
            
            wb.save(output_dir / name)

class EdgeCaseGenerator(BaseGenerator):
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]
        
        for name in output_names:
            wb = openpyxl.Workbook()
            # Remove default sheet
            default_ws = wb.active
            wb.remove(default_ws)
            
            # Empty Sheet
            wb.create_sheet("Empty")
            
            # Values Only
            ws_val = wb.create_sheet("ValuesOnly")
            for r in range(1, 11):
                for c in range(1, 11):
                    ws_val.cell(row=r, column=c, value=r*c)
            
            # Formulas Only
            ws_form = wb.create_sheet("FormulasOnly")
            for r in range(1, 11):
                for c in range(1, 11):
                    # Reference ValuesOnly sheet
                    col_letter = get_column_letter(c)
                    ws_form.cell(row=r, column=c, value=f"=ValuesOnly!{col_letter}{r}")
            
            wb.save(output_dir / name)

class AddressSanityGenerator(BaseGenerator):
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]
            
        targets = self.args.get('targets', ["A1", "B2", "Z10"])
        
        for name in output_names:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Addresses"
            
            for addr in targets:
                ws[addr] = addr
                
            wb.save(output_dir / name)

class ValueFormulaGenerator(BaseGenerator):
    """PG3: Types, formulas, values"""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]
            
        for name in output_names:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Types"
            
            ws['A1'] = 42
            ws['A2'] = "hello"
            ws['A3'] = True
            # A4 empty
            
            ws['B1'] = "=A1+1"
            ws['B2'] = '="hello" & " world"'
            ws['B3'] = "=A1>0"
            
            output_path = output_dir / name
            wb.save(output_path)
            self._inject_formula_caches(output_path)

    def _inject_formula_caches(self, path: Path):
        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        with zipfile.ZipFile(path, "r") as zf:
            sheet_xml = zf.read("xl/worksheets/sheet1.xml")
            other_files = {
                info.filename: zf.read(info.filename)
                for info in zf.infolist()
                if info.filename != "xl/worksheets/sheet1.xml"
            }

        root = ET.fromstring(sheet_xml)

        def update_cell(ref: str, value: str, cell_type: str | None = None):
            cell = root.find(f".//{{{ns}}}c[@r='{ref}']")
            if cell is None:
                return
            if cell_type:
                cell.set("t", cell_type)
            v = cell.find(f"{{{ns}}}v")
            if v is None:
                v = ET.SubElement(cell, f"{{{ns}}}v")
            v.text = value

        update_cell("B1", "43")
        update_cell("B2", "hello world", "str")
        update_cell("B3", "1", "b")

        ET.register_namespace("", ns)
        updated_sheet = ET.tostring(root, encoding="utf-8", xml_declaration=False)
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("xl/worksheets/sheet1.xml", updated_sheet)
            for name, data in other_files.items():
                zf.writestr(name, data)


class FormulaSemanticPairGenerator(BaseGenerator):
    """Pair fixture: exercise formula_diff (formatting-only vs semantic change).

    Notes:
    - openpyxl does not compute cached formula values, so we inject <v> caches for determinism.
    - Keep it tiny so UI scenarios stay fast and deterministic.
    """

    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("formula_semantic_pair generator expects exactly two output filenames")

        def build_one(path: Path, variant: str):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            ws["A1"] = 10
            ws["A2"] = 5
            ws["B1"] = 2

            if variant == "a":
                # B changes only whitespace: should classify as FormattingOnly when semantic formula
                # diff is enabled.
                ws["C1"] = "=A1+A2"
                # B changes semantics.
                ws["C2"] = "=A1*B1"
            else:
                ws["C1"] = "=A1 + A2"
                ws["C2"] = "=A1*(B1+1)"

            wb.save(path)
            self._inject_formula_caches(path, variant)

        build_one(output_dir / output_names[0], "a")
        build_one(output_dir / output_names[1], "b")

    def _inject_formula_caches(self, path: Path, variant: str):
        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        with zipfile.ZipFile(path, "r") as zf:
            sheet_xml = zf.read("xl/worksheets/sheet1.xml")
            other_files = {
                info.filename: zf.read(info.filename)
                for info in zf.infolist()
                if info.filename != "xl/worksheets/sheet1.xml"
            }

        root = ET.fromstring(sheet_xml)

        def update_cell(ref: str, value: str):
            cell = root.find(f".//{{{ns}}}c[@r='{ref}']")
            if cell is None:
                return
            v = cell.find(f"{{{ns}}}v")
            if v is None:
                v = ET.SubElement(cell, f"{{{ns}}}v")
            v.text = value

        update_cell("C1", "15")
        update_cell("C2", "20" if variant == "a" else "30")

        ET.register_namespace("", ns)
        updated_sheet = ET.tostring(root, encoding="utf-8", xml_declaration=False)
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("xl/worksheets/sheet1.xml", updated_sheet)
            for name, data in other_files.items():
                zf.writestr(name, data)

class SingleCellDiffGenerator(BaseGenerator):
    """Generates a tiny pair of workbooks with a single differing cell."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("single_cell_diff generator expects exactly two output filenames")

        rows = self.args.get('rows', 3)
        cols = self.args.get('cols', 3)
        sheet = self.args.get('sheet', "Sheet1")
        target_cell = self.args.get('target_cell', "C3")
        value_a = self.args.get('value_a', "1")
        value_b = self.args.get('value_b', "2")

        def create_workbook(value, name: str):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet

            for r in range(1, rows + 1):
                for c in range(1, cols + 1):
                    ws.cell(row=r, column=c, value=f"R{r}C{c}")

            ws[target_cell] = value
            wb.save(output_dir / name)

        create_workbook(value_a, output_names[0])
        create_workbook(value_b, output_names[1])

class MultiCellDiffGenerator(BaseGenerator):
    """Generates workbook pairs that differ in multiple scattered cells."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("multi_cell_diff generator expects exactly two output filenames")

        rows = self.args.get("rows", 20)
        cols = self.args.get("cols", 10)
        sheet = self.args.get("sheet", "Sheet1")
        edits: List[Dict[str, Any]] = self.args.get("edits", [])

        self._create_workbook(output_dir / output_names[0], sheet, rows, cols, edits, "a")
        self._create_workbook(output_dir / output_names[1], sheet, rows, cols, edits, "b")

    def _create_workbook(
        self,
        path: Path,
        sheet: str,
        rows: int,
        cols: int,
        edits: List[Dict[str, Any]],
        value_key: str,
    ):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        self._fill_base_grid(ws, rows, cols)
        self._apply_edits(ws, edits, value_key)

        wb.save(path)

    def _fill_base_grid(self, ws, rows: int, cols: int):
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c, value=f"R{r}C{c}")

    def _apply_edits(self, ws, edits: List[Dict[str, Any]], value_key: str):
        value_field = f"value_{value_key}"

        for edit in edits:
            addr = edit.get("addr")
            if not addr:
                raise ValueError("multi_cell_diff edits require 'addr'")
            if value_field not in edit:
                raise ValueError(f"multi_cell_diff edits require '{value_field}'")
            ws[addr] = edit[value_field]

class GridTailDiffGenerator(BaseGenerator):
    """Generates workbook pairs for simple row/column tail append/delete scenarios."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("grid_tail_diff generator expects exactly two output filenames")

        mode = self.args.get("mode")
        sheet = self.args.get("sheet", "Sheet1")

        if mode == "row_append_bottom":
            self._row_append_bottom(output_dir, output_names, sheet)
        elif mode == "row_delete_bottom":
            self._row_delete_bottom(output_dir, output_names, sheet)
        elif mode == "col_append_right":
            self._col_append_right(output_dir, output_names, sheet)
        elif mode == "col_delete_right":
            self._col_delete_right(output_dir, output_names, sheet)
        else:
            raise ValueError(f"Unsupported grid_tail_diff mode: {mode}")

    def _row_append_bottom(self, output_dir: Path, output_names: List[str], sheet: str):
        base_rows = self.args.get("base_rows", 10)
        tail_rows = self.args.get("tail_rows", 2)
        cols = self.args.get("cols", 3)

        self._write_rows(output_dir / output_names[0], sheet, base_rows, cols, 1)
        self._write_rows(
            output_dir / output_names[1],
            sheet,
            base_rows + tail_rows,
            cols,
            1,
        )

    def _row_delete_bottom(self, output_dir: Path, output_names: List[str], sheet: str):
        base_rows = self.args.get("base_rows", 10)
        tail_rows = self.args.get("tail_rows", 2)
        cols = self.args.get("cols", 3)

        self._write_rows(
            output_dir / output_names[0],
            sheet,
            base_rows + tail_rows,
            cols,
            1,
        )
        self._write_rows(output_dir / output_names[1], sheet, base_rows, cols, 1)

    def _col_append_right(self, output_dir: Path, output_names: List[str], sheet: str):
        base_cols = self.args.get("base_cols", 4)
        tail_cols = self.args.get("tail_cols", 2)
        rows = self.args.get("rows", 5)

        self._write_cols(output_dir / output_names[0], sheet, rows, base_cols)
        self._write_cols(
            output_dir / output_names[1],
            sheet,
            rows,
            base_cols + tail_cols,
        )

    def _col_delete_right(self, output_dir: Path, output_names: List[str], sheet: str):
        base_cols = self.args.get("base_cols", 4)
        tail_cols = self.args.get("tail_cols", 2)
        rows = self.args.get("rows", 5)

        self._write_cols(
            output_dir / output_names[0],
            sheet,
            rows,
            base_cols + tail_cols,
        )
        self._write_cols(output_dir / output_names[1], sheet, rows, base_cols)

    def _write_rows(self, path: Path, sheet: str, rows: int, cols: int, start_value: int):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        for r in range(1, rows + 1):
            ws.cell(row=r, column=1, value=start_value + r - 1)
            for c in range(2, cols + 1):
                ws.cell(row=r, column=c, value=f"R{r}C{c}")

        wb.save(path)

    def _write_cols(self, path: Path, sheet: str, rows: int, cols: int):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c, value=f"R{r}C{c}")

        wb.save(path)

class RowAlignmentG8Generator(BaseGenerator):
    """Generates workbook pairs for G8-style middle row insert/delete scenarios."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("row_alignment_g8 generator expects exactly two output filenames")

        mode = self.args.get("mode")
        sheet = self.args.get("sheet", "Sheet1")
        base_rows = self.args.get("base_rows", 10)
        cols = self.args.get("cols", 5)
        insert_at = self.args.get("insert_at", 6)  # 1-based position in B
        delete_row = self.args.get("delete_row", 6)  # 1-based position in A
        edit_row = self.args.get("edit_row")  # Optional extra edit row (1-based in B after insert)
        edit_col = self.args.get("edit_col", 2)  # 1-based column for extra edit

        base_data = [self._base_row_values(idx, cols) for idx in range(1, base_rows + 1)]

        if mode == "insert":
            data_a = base_data
            data_b = self._with_insert(base_data, insert_at, cols)
        elif mode == "delete":
            data_a = base_data
            data_b = self._with_delete(base_data, delete_row)
        elif mode == "insert_with_edit":
            data_a = base_data
            data_b = self._with_insert(base_data, insert_at, cols)
            target_row = edit_row or (insert_at + 2)
            if 1 <= target_row <= len(data_b):
                row_values = list(data_b[target_row - 1])
                col_index = max(1, min(edit_col, cols)) - 1
                row_values[col_index] = "EditedAfterInsert"
                data_b[target_row - 1] = row_values
        else:
            raise ValueError(f"Unsupported row_alignment_g8 mode: {mode}")

        self._write_workbook(output_dir / output_names[0], sheet, data_a)
        self._write_workbook(output_dir / output_names[1], sheet, data_b)

    def _base_row_values(self, row_number: int, cols: int) -> List[str]:
        return [f"Row{row_number}_Col{c}" for c in range(1, cols + 1)]

    def _insert_row_values(self, cols: int) -> List[str]:
        return [f"Inserted_Row_Col{c}" for c in range(1, cols + 1)]

    def _with_insert(self, base_data: List[List[str]], insert_at: int, cols: int) -> List[List[str]]:
        insert_idx = max(1, min(insert_at, len(base_data) + 1))
        insert_row = self._insert_row_values(cols)
        return base_data[: insert_idx - 1] + [insert_row] + base_data[insert_idx - 1 :]

    def _with_delete(self, base_data: List[List[str]], delete_row: int) -> List[List[str]]:
        if not (1 <= delete_row <= len(base_data)):
            raise ValueError(f"delete_row must be within 1..{len(base_data)}")
        return base_data[: delete_row - 1] + base_data[delete_row:]

    def _write_workbook(self, path: Path, sheet: str, rows: List[List[str]]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        for r_idx, row_values in enumerate(rows, start=1):
            for c_idx, value in enumerate(row_values, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(path)

class RowAlignmentG10Generator(BaseGenerator):
    """Generates workbook pairs for G10 contiguous row block insert/delete scenarios."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("row_alignment_g10 generator expects exactly two output filenames")

        mode = self.args.get("mode")
        sheet = self.args.get("sheet", "Sheet1")
        base_rows = self.args.get("base_rows", 10)
        cols = self.args.get("cols", 5)
        block_rows = self.args.get("block_rows", 4)
        insert_at = self.args.get("insert_at", 4)  # 1-based position of first inserted row in B
        delete_start = self.args.get("delete_start", 4)  # 1-based starting row in A to delete

        base_data = [self._row_values(idx, cols, 0) for idx in range(1, base_rows + 1)]

        if mode == "block_insert":
            data_a = base_data
            data_b = self._with_block_insert(base_data, insert_at, block_rows, cols)
        elif mode == "block_delete":
            data_a = base_data
            data_b = self._with_block_delete(base_data, delete_start, block_rows)
        else:
            raise ValueError(f"Unsupported row_alignment_g10 mode: {mode}")

        self._write_workbook(output_dir / output_names[0], sheet, data_a)
        self._write_workbook(output_dir / output_names[1], sheet, data_b)

    def _row_values(self, row_number: int, cols: int, offset: int) -> List[int]:
        row_id = row_number + offset
        values = [row_id]
        for c in range(1, cols):
            values.append(row_id * 10 + c)
        return values

    def _block_rows(self, count: int, cols: int) -> List[List[int]]:
        return [self._row_values(1000 + idx, cols, 0) for idx in range(1, count + 1)]

    def _with_block_insert(
        self, base_data: List[List[int]], insert_at: int, block_rows: int, cols: int
    ) -> List[List[int]]:
        insert_idx = max(1, min(insert_at, len(base_data) + 1)) - 1
        block = self._block_rows(block_rows, cols)
        return base_data[:insert_idx] + block + base_data[insert_idx:]

    def _with_block_delete(
        self, base_data: List[List[int]], delete_start: int, block_rows: int
    ) -> List[List[int]]:
        if not (1 <= delete_start <= len(base_data)):
            raise ValueError(f"delete_start must be within 1..{len(base_data)}")
        if delete_start - 1 + block_rows > len(base_data):
            raise ValueError("delete block exceeds base data length")

        delete_idx = delete_start - 1
        return base_data[:delete_idx] + base_data[delete_idx + block_rows :]

    def _write_workbook(self, path: Path, sheet: str, rows: List[List[int]]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        for r_idx, row_values in enumerate(rows, start=1):
            for c_idx, value in enumerate(row_values, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(path)

class RowBlockMoveG11Generator(BaseGenerator):
    """Generates workbook pairs for G11 exact row block move scenarios."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("row_block_move_g11 generator expects exactly two output filenames")

        sheet = self.args.get("sheet", "Sheet1")
        total_rows = self.args.get("total_rows", 20)
        cols = self.args.get("cols", 5)
        block_rows = self.args.get("block_rows", 4)
        src_start = self.args.get("src_start", 5)
        dst_start = self.args.get("dst_start", 13)

        if block_rows <= 0:
            raise ValueError("block_rows must be positive")
        if src_start < 1 or src_start + block_rows - 1 > total_rows:
            raise ValueError("source block must fit within total_rows")
        if dst_start < 1 or dst_start + block_rows - 1 > total_rows:
            raise ValueError("destination block must fit within total_rows")

        src_end = src_start + block_rows - 1
        dst_end = dst_start + block_rows - 1
        if not (src_end < dst_start or dst_end < src_start):
            raise ValueError("source and destination blocks must not overlap")

        rows_a = self._build_rows(total_rows, cols, src_start, block_rows)
        rows_b = self._move_block(rows_a, src_start, block_rows, dst_start)

        self._write_workbook(output_dir / output_names[0], sheet, rows_a)
        self._write_workbook(output_dir / output_names[1], sheet, rows_b)

    def _build_rows(self, total_rows: int, cols: int, src_start: int, block_rows: int) -> List[List[str]]:
        block_end = src_start + block_rows - 1
        rows: List[List[str]] = []
        for r in range(1, total_rows + 1):
            if src_start <= r <= block_end:
                rows.append([f"BLOCK_r{r}_c{c}" for c in range(1, cols + 1)])
            else:
                rows.append([f"R{r}_C{c}" for c in range(1, cols + 1)])
        return rows

    def _move_block(
        self, rows: List[List[str]], src_start: int, block_rows: int, dst_start: int
    ) -> List[List[str]]:
        rows_b = [list(r) for r in rows]
        src_idx = src_start - 1
        src_end = src_idx + block_rows
        block = rows_b[src_idx:src_end]
        del rows_b[src_idx:src_end]

        dst_idx = min(dst_start - 1, len(rows_b))

        rows_b[dst_idx:dst_idx] = block
        return rows_b

    def _write_workbook(self, path: Path, sheet: str, rows: List[List[str]]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        for r_idx, row_values in enumerate(rows, start=1):
            for c_idx, value in enumerate(row_values, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(path)

class RowFuzzyMoveG13Generator(BaseGenerator):
    """Generates workbook pairs for G13 fuzzy row block move scenarios with internal edits."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("row_fuzzy_move_g13 generator expects exactly two output filenames")

        sheet = self.args.get("sheet", "Data")
        total_rows = self.args.get("total_rows", 24)
        cols = self.args.get("cols", 6)
        block_rows = self.args.get("block_rows", 4)
        src_start = self.args.get("src_start", 5)
        dst_start = self.args.get("dst_start", 14)
        edits = self.args.get(
            "edits",
            [
                {"row_offset": 1, "col": 3, "delta": 1},
            ],
        )

        if block_rows <= 0:
            raise ValueError("block_rows must be positive")
        if src_start < 1 or src_start + block_rows - 1 > total_rows:
            raise ValueError("source block must fit within total_rows")
        if dst_start < 1 or dst_start + block_rows - 1 > total_rows:
            raise ValueError("destination block must fit within total_rows")

        src_end = src_start + block_rows - 1
        dst_end = dst_start + block_rows - 1
        if not (src_end < dst_start or dst_end < src_start):
            raise ValueError("source and destination blocks must not overlap")

        rows_a = self._build_rows(total_rows, cols, src_start, block_rows)
        rows_b = self._move_block(rows_a, src_start, block_rows, dst_start)
        self._apply_edits(rows_b, dst_start, block_rows, cols, edits)

        self._write_workbook(output_dir / output_names[0], sheet, rows_a)
        self._write_workbook(output_dir / output_names[1], sheet, rows_b)

    def _build_rows(self, total_rows: int, cols: int, block_start: int, block_rows: int) -> List[List[int]]:
        block_end = block_start + block_rows - 1
        rows: List[List[int]] = []
        for r in range(1, total_rows + 1):
            if block_start <= r <= block_end:
                row_id = 1_000 + (r - block_start)
            else:
                row_id = r
            row_values = [row_id]
            for c in range(1, cols):
                row_values.append(row_id * 10 + c)
            rows.append(row_values)
        return rows

    def _move_block(
        self, rows: List[List[int]], src_start: int, block_rows: int, dst_start: int
    ) -> List[List[int]]:
        rows_b = [list(r) for r in rows]
        src_idx = src_start - 1
        src_end = src_idx + block_rows
        block = rows_b[src_idx:src_end]
        del rows_b[src_idx:src_end]

        dst_idx = min(dst_start - 1, len(rows_b))
        rows_b[dst_idx:dst_idx] = block
        return rows_b

    def _apply_edits(
        self,
        rows: List[List[int]],
        dst_start: int,
        block_rows: int,
        cols: int,
        edits: List[Dict[str, Any]],
    ):
        dst_idx = dst_start - 1
        if dst_idx + block_rows > len(rows):
            return

        for edit in edits:
            row_offset = int(edit.get("row_offset", 0))
            col = int(edit.get("col", 1))
            delta = int(edit.get("delta", 1))

            if row_offset < 0 or row_offset >= block_rows:
                continue

            col_idx = max(1, min(col, cols)) - 1
            target_row = dst_idx + row_offset
            if col_idx >= len(rows[target_row]):
                continue
            rows[target_row][col_idx] += delta

    def _write_workbook(self, path: Path, sheet: str, rows: List[List[int]]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        for r_idx, row_values in enumerate(rows, start=1):
            for c_idx, value in enumerate(row_values, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(path)

class ColumnMoveG12Generator(BaseGenerator):
    """Generates workbook pairs for G12 exact column move scenarios."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("column_move_g12 generator expects exactly two output filenames")

        sheet = self.args.get("sheet", "Data")
        cols = self.args.get("cols", 8)
        data_rows = self.args.get("data_rows", 9)
        src_col = self.args.get("src_col", 3)
        dst_col = self.args.get("dst_col", 6)

        if not (1 <= src_col <= cols):
            raise ValueError("src_col must be within 1..cols")
        if not (1 <= dst_col <= cols):
            raise ValueError("dst_col must be within 1..cols")
        if src_col == dst_col:
            raise ValueError("src_col and dst_col must differ for a move")

        base_rows = self._build_rows(cols, data_rows, src_col)
        moved_rows = self._move_column(base_rows, src_col, dst_col)

        self._write_workbook(output_dir / output_names[0], sheet, base_rows)
        self._write_workbook(output_dir / output_names[1], sheet, moved_rows)

    def _build_rows(self, cols: int, data_rows: int, key_col: int) -> List[List[Any]]:
        header: List[Any] = []
        for c in range(1, cols + 1):
            if c == key_col:
                header.append("C_key")
            else:
                header.append(f"Col{c}")

        rows: List[List[Any]] = [header]
        for r in range(1, data_rows + 1):
            row: List[Any] = []
            for c in range(1, cols + 1):
                if c == key_col:
                    row.append(100 * r)
                else:
                    row.append(r * 10 + c)
            rows.append(row)

        return rows

    def _move_column(
        self, rows: List[List[Any]], src_col: int, dst_col: int
    ) -> List[List[Any]]:
        src_idx = src_col - 1
        dst_idx = dst_col - 1
        moved_rows: List[List[Any]] = []

        for row in rows:
            new_row = list(row)
            value = new_row.pop(src_idx)
            insert_at = max(0, min(dst_idx, len(new_row)))
            new_row.insert(insert_at, value)
            moved_rows.append(new_row)

        return moved_rows

    def _write_workbook(self, path: Path, sheet: str, rows: List[List[Any]]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        for r_idx, row_values in enumerate(rows, start=1):
            for c_idx, value in enumerate(row_values, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(path)

class RectBlockMoveG12Generator(BaseGenerator):
    """Generates workbook pairs for G12 exact rectangular block move scenarios."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("rect_block_move_g12 generator expects exactly two output filenames")

        sheet = self.args.get("sheet", "Data")
        rows = self.args.get("rows", 15)
        cols = self.args.get("cols", 15)
        src_top = self.args.get("src_top", 3)  # 1-based
        src_left = self.args.get("src_left", 2)  # 1-based (column B)
        dst_top = self.args.get("dst_top", 10)  # 1-based
        dst_left = self.args.get("dst_left", 7)  # 1-based (column G)
        block_rows = self.args.get("block_rows", 3)
        block_cols = self.args.get("block_cols", 3)

        self._write_workbook(
            output_dir / output_names[0],
            sheet,
            rows,
            cols,
            src_top,
            src_left,
            block_rows,
            block_cols,
        )
        self._write_workbook(
            output_dir / output_names[1],
            sheet,
            rows,
            cols,
            dst_top,
            dst_left,
            block_rows,
            block_cols,
        )

    def _write_workbook(
        self,
        path: Path,
        sheet: str,
        rows: int,
        cols: int,
        block_top: int,
        block_left: int,
        block_rows: int,
        block_cols: int,
    ):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        self._fill_background(ws, rows, cols)
        self._write_block(ws, block_top, block_left, block_rows, block_cols)

        wb.save(path)

    def _fill_background(self, ws, rows: int, cols: int):
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c, value=self._background_value(r, c))

    def _background_value(self, row: int, col: int) -> int:
        return 1000 * row + col

    def _write_block(self, ws, top: int, left: int, block_rows: int, block_cols: int):
        for r_offset in range(block_rows):
            for c_offset in range(block_cols):
                value = 9000 + r_offset * 10 + c_offset
                ws.cell(row=top + r_offset, column=left + c_offset, value=value)

class ColumnAlignmentG9Generator(BaseGenerator):
    """Generates workbook pairs for G9-style middle column insert/delete scenarios."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("column_alignment_g9 generator expects exactly two output filenames")

        mode = self.args.get("mode")
        sheet = self.args.get("sheet", "Data")
        base_cols = self.args.get("cols", 8)
        data_rows = self.args.get("data_rows", 9)  # excludes header
        insert_at = self.args.get("insert_at", 4)  # 1-based position in B after insert
        delete_col = self.args.get("delete_col", 4)
        edit_row = self.args.get("edit_row", 8)
        edit_col_after_insert = self.args.get("edit_col_after_insert", 7)

        base_table = self._base_table(base_cols, data_rows)

        if mode == "insert":
            data_a = self._clone_rows(base_table)
            data_b = self._with_insert(base_table, insert_at)
        elif mode == "delete":
            data_a = self._clone_rows(base_table)
            data_b = self._with_delete(base_table, delete_col)
        elif mode == "insert_with_edit":
            data_a = self._clone_rows(base_table)
            data_b = self._with_insert(base_table, insert_at)
            row_idx = max(2, min(edit_row, len(data_b))) - 1  # stay below header
            col_idx = max(1, min(edit_col_after_insert, len(data_b[row_idx]))) - 1
            data_b[row_idx][col_idx] = "EditedAfterInsert"
        else:
            raise ValueError(f"Unsupported column_alignment_g9 mode: {mode}")

        self._write_workbook(output_dir / output_names[0], sheet, data_a)
        self._write_workbook(output_dir / output_names[1], sheet, data_b)

    def _base_table(self, cols: int, data_rows: int) -> List[List[str]]:
        header = [f"Col{c}" for c in range(1, cols + 1)]
        rows = [header]
        for r in range(1, data_rows + 1):
            rows.append([f"R{r}_C{c}" for c in range(1, cols + 1)])
        return rows

    def _with_insert(self, base_data: List[List[str]], insert_at: int) -> List[List[str]]:
        insert_idx = max(1, min(insert_at, len(base_data[0]) + 1))
        result: List[List[str]] = []
        for row_idx, row in enumerate(base_data):
            new_row = list(row)
            value = "Inserted" if row_idx == 0 else f"Inserted_{row_idx}"
            new_row.insert(insert_idx - 1, value)
            result.append(new_row)
        return result

    def _with_delete(self, base_data: List[List[str]], delete_col: int) -> List[List[str]]:
        if not base_data:
            return []
        if not (1 <= delete_col <= len(base_data[0])):
            raise ValueError(f"delete_col must be within 1..{len(base_data[0])}")
        result: List[List[str]] = []
        for row in base_data:
            new_row = list(row)
            del new_row[delete_col - 1]
            result.append(new_row)
        return result

    def _clone_rows(self, rows: List[List[str]]) -> List[List[str]]:
        return [list(r) for r in rows]

    def _write_workbook(self, path: Path, sheet: str, rows: List[List[str]]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet

        for r_idx, row_values in enumerate(rows, start=1):
            for c_idx, value in enumerate(row_values, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(path)

class SheetCaseRenameGenerator(BaseGenerator):
    """Generates a pair of workbooks that differ only by sheet name casing, with optional cell edit."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("sheet_case_rename generator expects exactly two output filenames")

        sheet_a = self.args.get("sheet_a", "Sheet1")
        sheet_b = self.args.get("sheet_b", "sheet1")
        cell = self.args.get("cell", "A1")
        value_a = self.args.get("value_a", 1.0)
        value_b = self.args.get("value_b", value_a)

        def create_workbook(sheet_name: str, value, output_name: str):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws[cell] = value
            wb.save(output_dir / output_name)

        create_workbook(sheet_a, value_a, output_names[0])
        create_workbook(sheet_b, value_b, output_names[1])

class Pg6SheetScenarioGenerator(BaseGenerator):
    """Generates workbook pairs for PG6 sheet add/remove/rename vs grid responsibilities."""
    def generate(self, output_dir: Path, output_names: Union[str, List[str]]):
        if isinstance(output_names, str):
            output_names = [output_names]

        if len(output_names) != 2:
            raise ValueError("pg6_sheet_scenario generator expects exactly two output filenames")

        mode = self.args.get("mode")
        a_path = output_dir / output_names[0]
        b_path = output_dir / output_names[1]

        if mode == "sheet_added":
            self._gen_sheet_added(a_path, b_path)
        elif mode == "sheet_removed":
            self._gen_sheet_removed(a_path, b_path)
        elif mode == "sheet_renamed":
            self._gen_sheet_renamed(a_path, b_path)
        elif mode == "sheet_and_grid_change":
            self._gen_sheet_and_grid_change(a_path, b_path)
        else:
            raise ValueError(f"Unsupported PG6 mode: {mode}")

    def _fill_grid(self, worksheet, rows: int, cols: int, prefix: str = "R"):
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                worksheet.cell(row=r, column=c, value=f"{prefix}{r}C{c}")

    def _gen_sheet_added(self, a_path: Path, b_path: Path):
        wb_a = openpyxl.Workbook()
        ws_main_a = wb_a.active
        ws_main_a.title = "Main"
        self._fill_grid(ws_main_a, 5, 5)
        wb_a.save(a_path)

        wb_b = openpyxl.Workbook()
        ws_main_b = wb_b.active
        ws_main_b.title = "Main"
        self._fill_grid(ws_main_b, 5, 5)
        ws_new = wb_b.create_sheet("NewSheet")
        self._fill_grid(ws_new, 3, 3, prefix="N")
        wb_b.save(b_path)

    def _gen_sheet_removed(self, a_path: Path, b_path: Path):
        wb_a = openpyxl.Workbook()
        ws_main_a = wb_a.active
        ws_main_a.title = "Main"
        self._fill_grid(ws_main_a, 5, 5)
        ws_old = wb_a.create_sheet("OldSheet")
        self._fill_grid(ws_old, 3, 3, prefix="O")
        wb_a.save(a_path)

        wb_b = openpyxl.Workbook()
        ws_main_b = wb_b.active
        ws_main_b.title = "Main"
        self._fill_grid(ws_main_b, 5, 5)
        wb_b.save(b_path)

    def _gen_sheet_renamed(self, a_path: Path, b_path: Path):
        wb_a = openpyxl.Workbook()
        ws_old = wb_a.active
        ws_old.title = "OldName"
        self._fill_grid(ws_old, 3, 3)
        wb_a.save(a_path)

        wb_b = openpyxl.Workbook()
        ws_new = wb_b.active
        ws_new.title = "NewName"
        self._fill_grid(ws_new, 3, 3)
        wb_b.save(b_path)

    def _gen_sheet_and_grid_change(self, a_path: Path, b_path: Path):
        base_rows = 5
        base_cols = 5

        wb_a = openpyxl.Workbook()
        ws_main_a = wb_a.active
        ws_main_a.title = "Main"
        self._fill_grid(ws_main_a, base_rows, base_cols)
        ws_aux_a = wb_a.create_sheet("Aux")
        self._fill_grid(ws_aux_a, 3, 3, prefix="A")
        wb_a.save(a_path)

        wb_b = openpyxl.Workbook()
        ws_main_b = wb_b.active
        ws_main_b.title = "Main"
        self._fill_grid(ws_main_b, base_rows, base_cols)
        ws_main_b["A1"] = "Main changed 1"
        ws_main_b["B2"] = "Main changed 2"
        ws_main_b["C3"] = "Main changed 3"

        ws_aux_b = wb_b.create_sheet("Aux")
        self._fill_grid(ws_aux_b, 3, 3, prefix="A")

        ws_scratch = wb_b.create_sheet("Scratch")
        self._fill_grid(ws_scratch, 2, 2, prefix="S")
        wb_b.save(b_path)
